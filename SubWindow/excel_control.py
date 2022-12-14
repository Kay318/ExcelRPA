from csv import excel
import xlwings as xw
from datetime import datetime
from openpyxl.styles import Alignment
from openpyxl.styles import Border, Side
from openpyxl.styles import PatternFill
from openpyxl.styles import borders
from openpyxl.styles.fonts import Font
from typing import List
import openpyxl as xl
import win32com.client as win32
from openpyxl.drawing.image import Image
from openpyxl.worksheet.worksheet import Worksheet
from Settings import Setup as sp
import os
import sys
import string
import time
from PyQt5.QtCore import *
from PyQt5.QtGui import *
from PyQt5.QtWidgets import *
from pathlib import Path
from Log import LogManager

sys.path.append(str(Path(__file__).parents[1]))
from DataBase import DB as db

class ExcelRun(QThread):
    progressBarValue = pyqtSignal(int)
    signal_done = pyqtSignal(int)
    
    def __init__(self, parent=None):
        super().__init__(parent)
        self.parent = parent
        self.path = parent.path
        self.lang_List = parent.selected_langList
        self.testBool = parent.testBool
        self.path_file = parent.path_file
        
        self.IMG_SHEET_HEIGHTSIZE = 115 # 이미지 간격
        self.IMG_FAINAL_WIDTH = 50
        self.SHEET_WIDTH_SHORTSIZE = 15 # 시트 열 기본 작은 크기
        self.TABLE_CELL_COLOR = 43 # 테이블 컬러
        self.SHEET_EvaluationListSIZE = None # 평가목록 넓이
        
        self.set_row = 1
        self.start_column = 1
        self.lang_cnt = 1
        self.sp = sp.Settings()
        self.testList, _ = self.sp.read_setup(table = "Test_List")
        # self.get_lang_List = self.lang_List
        
        self.column = []
        self.excel_setup()
        
        for i in range(65, 91):
            self.column.append(chr(i))
        
    def run(self):

        if self.testBool:
            self.fail_checkList = ["FAIL", "N/A", "N/T", ""]
            self.wb = xl.Workbook()
            
            ws = self.wb.active
            ws.title = "SUMMARY"
            self.history_rows = 1
            
            for idx, lang in enumerate(self.lang_List):
                self.start_percent = idx/len(self.lang_List)
                self.split_percent = 1/len(self.lang_List)
    
                self.create_sheet_history(lang=lang, ws= ws)

                active = self.wb.create_sheet(title=lang)

                self.excel_data_input(active = active, lang = lang)

            self.wb.save(self.path_file)
            self.progressBarValue.emit(100)
            self.signal_done.emit(1)
        else:
            try:
                app = xw.App(visible=False, add_book=False)
                app.display_alerts=False
                wb = app.books.open(self.path_file)
            
                self.ws_summary = wb.sheets('SUMMARY')
                sheets = wb.sheets
                sheets_li = [s.name for s in sheets]
                
                for idx, lang in enumerate(self.lang_List):
                    self.start_percent = idx/len(self.lang_List)
                    self.split_percent = 1/len(self.lang_List)

                    # 시트 있으면 삭제 후 생성
                    if lang in sheets_li:
                        wb.sheets(lang).delete()
                    self.ws2 = wb.sheets.add(lang)
                    
                    # DB에서 필요한 데이터 불러오기
                    self.select_DB(lang)

                    self.insert_FirstLine()
                    self.insert_langSheetData()
                    self.set_langSheetStyle()
                    
                    self.ROW_NUM = 1
                    
                    while True:
                        summary_languageList = self.ws_summary.range(f"A{self.ROW_NUM}").expand("down").value
                        if not bool(summary_languageList):
                            if bool(self.summaryData):
                                self.insert_summary(lang)
                            break
                        
                        if lang in summary_languageList:
                            self.ws_summary.range(f'{self.ROW_NUM}:{self.ROW_NUM+len(summary_languageList)}').delete()
                            self.insert_summary(lang)
                            break
                                
                        self.ROW_NUM = self.ROW_NUM + len(summary_languageList) + 1

                    # 기존 엑셀의 시트 순서대로 배치, 없는 시트는 맨마지막에 배치
                    if lang in sheets_li:
                        self.ws2.api.Move(Before=sheets[sheets_li.index(lang)].api, After=None)
                    else:
                        self.ws2.api.Move(Before=None, After=sheets[len(sheets_li)].api)
                    
                # SUMMARY 시트를 맨앞으로 당겨감
                self.ws_summary.api.Move(Before=sheets[0].api, After=None)
                htime = datetime.now().strftime("%y%m%d_%H%M%S")
                fileName = f"{htime}_다국어자동화.xlsx"
                savePath = os.path.join(self.path, fileName)
                wb.save(savePath)
                self.progressBarValue.emit(100)
                
            except Exception as e:
                LogManager.HLOG.info(f"기존 엑셀 편집 중 오류", e)
                QMessageBox.warning(self.parent, '주의', '엑셀 편집이 실패되었습니다.\n파일 끄고 다시 해주세요.')
            finally:
                wb.close()
                app.quit()
                self.signal_done.emit(1)
            
            # self.progressBarValue.emit(100)
            # self.signal_done.emit(1)





            # TITLE_TERGET = 1
            # del_lastCell = False
            # self.fail_checkList = ["FAIL", "N/A", "N/T", " ", None, ""]

            # excel = win32.Dispatch("Excel.Application")
            # excel.Visible = False
            # excel.DisplayAlerts=False
            # self.wb = excel.Workbooks.Open(self.path_file)
            # print(f'self.lang_List : {self.lang_List}')
            # for jdx, lang in enumerate(self.lang_List):
            #     self.historyUpdate_rows = 2
            #     self.start_percent = jdx/len(self.lang_List)
            #     self.split_percent = 1/len(self.lang_List)

            #     print(f'self.start_percent : {self.start_percent}')
            #     print(f'self.split_percent : {self.split_percent}')
            #     if (lang == self.lang_List[len(self.lang_List) - 1]):
            #         del_lastCell = True

            #     self.update_sheet_history(lang, del_lastCell, TITLE_TERGET)
            #     ws = None
            #     try:
            #         ws = self.wb.Worksheets(f"{lang}")
            #         ws.Delete()
            #     except Exception as e:
            #         print(e)

            #     ws = self.wb.Worksheets.Add()
            #     ws.Name = lang

            #     kr_list = [] # 현재 엑셀 데이터
            #     idx = 1

            #     while(ws.Cells(TITLE_TERGET, idx).Value != None):
            #         value = "" if "" == ws.Cells(TITLE_TERGET, idx).Value else ws.Cells(TITLE_TERGET, idx).Value
                    
            #         kr_list.append(value)
            #         idx = idx + 1

            #     self.__equalsVerification__(kr_list=kr_list, ws = self.wb.Worksheets(lang), lang = lang, terget=TITLE_TERGET)
            #     QApplication.processEvents()
        
            # # 시트 순서 적용

            # print("Save:전")
            # self.wb.Save()
            # # 시트 전체 이름 조회 필요
            # self.wb.Worksheets("SUMMARY").Move(self.wb.Worksheets(1))
            # print("Save:전2")
            # self.wb.Save()
            # print("Save:후")
            # self.wb.Close(True)
            # print("wb close")
            # excel.Quit()
            # print("excel close")
            
            # self.progressBarValue.emit(100)
            # self.signal_done.emit(1)

    def insert_langSheetData(self):
        """언어별 시트에 데이터 입력하는 함수
        """

        self.summaryData = []
        for i, data in enumerate(self.dataList):
            # 데이터 입력
            self.ws2.range(f'A{i+2}').value=data

            # 이미지 삽입
            img_path = data[0].replace("/", "\\")
            if os.path.isfile(img_path):
                self.ws2.pictures.add(img_path, 
                                left=self.ws2.range(f"A{i+2}").left,
                                top=self.ws2.range(f"A{i+2}").top,
                                width=self.IMG_WIDTHSIZE,
                                height=self.IMG_HEIGHTSIZE)
            else:
                self.ws2[f'A{i+2}'].value=f"파일 없음\n{img_path}"

            # 이미지 셀 너비, 높이 설정
            self.ws2.range(f'A{i+2}').row_height = self.IMG_HEIGHTSIZE
            self.ws2.range(f'A{i+2}').column_width = self.IMG_FAINAL_WIDTH
                       
            # SUMMARY시트에 삽입할 데이터 저장
            if data[1:len(self.testList)+1].count('PASS') != len(self.testList):
                self.summaryData.append(data)

            percent_val = round((self.start_percent + ((i+1)/len(self.dataList))*self.split_percent)*100)
            if percent_val > 97:
                percent_val = 97
            self.progressBarValue.emit(percent_val)

    def select_DB(self, lang):
        """DB 조회하면서 컬럼명, 내용을 세팅하는 함수

        Args:
            lang (_type_): 현재 편집중인 언어
        """
        self.dataList = db.db_select(f"SELECT * FROM '{lang}'")
        self.sql_col_list = db.db_columns(f"SELECT * FROM '{lang}'")
        self.summary_col_list = self.sql_col_list[:]
        self.summary_col_list.insert(0, '언어')

    def insert_FirstLine(self):
        """언어별 타이틀 쓰는 함수
        """

        for i, col_name in enumerate(self.sql_col_list):
            self.ws2.cells(1, i+1).value = col_name
            if col_name in self.testList:
                self.ws2.cells(1, i+1).column_width = self.SHEET_EvaluationListSIZE
            else:
                self.ws2.cells(1, i+1).column_width = self.SHEET_WIDTHSIZE
                
        firstRange = self.ws2.range("A1").expand('right')
        firstRange.rows.autofit()
        firstRange.api.WrapText = True
        firstRange.color = xw.utils.rgb_to_int((153,204,000))       # 배경색: 초록색

    def setBorder(self, range):
        """전체 테두리 추가

        Args:
            range (_type_): 적용할 range
        """
        range.api.Borders.Weight = 2
        
    def alignCenter(self, range):
        """가운데 맞춤

        Args:
            range (_type_): 적용할 range
        """
        range.api.HorizontalAlignment = xw.constants.HAlign.xlHAlignCenter
        range.api.VerticalAlignment = xw.constants.VAlign.xlVAlignCenter

    def stop(self):
        self.terminate()

    def insert_summary(self, lang):
        """SUMMARY 입력하는 함수

        Args:
            lang (_type_): 편집 중인 언어
        """
        if bool(self.summaryData):
            START_ROW = self.ROW_NUM
            self.ws_summary.range(f'{self.ROW_NUM}:{self.ROW_NUM+len(self.summaryData)+1}').insert()
            
            self.insert_summaryTitle()
            self.ROW_NUM += 1
            self.insert_summaryData(lang)
            self.set_summaryStyle(START_ROW)

    def insert_summaryTitle(self):
        """SUMMARY 타이틀 입력하는 함수
           - 가운데 맞춤, 배경색, 자동 줄바꿈, 행높이 자동 설정 포함
        """
        for i, title in enumerate(self.summary_col_list):
            self.ws_summary.cells(self.ROW_NUM, i+1).value = title
            if title in self.testList:
                self.ws_summary.cells(self.ROW_NUM, i+1).column_width = self.SHEET_EvaluationListSIZE
            elif title == "언어":
                self.ws_summary.cells(self.ROW_NUM, i+1).column_width = 14
            else:
                self.ws_summary.cells(self.ROW_NUM, i+1).column_width = self.SHEET_WIDTHSIZE
                
        titleRange = self.ws_summary.range(f"A{self.ROW_NUM}").expand('right')
        titleRange.rows.autofit()
        titleRange.api.WrapText = True
        titleRange.color = xw.utils.rgb_to_int((153,204,000))       # 배경색: 초록색

    def insert_summaryData(self, lang):
        """SUMMARY 시트에서 선택한 언어에 대한 데이터 입력하는 함수

        Args:
            lang (_type_): 선택한 언어
        """
        for data in self.summaryData:
            self.ws_summary.cells(self.ROW_NUM, 1).value = lang
            self.ws_summary.range(f"B{self.ROW_NUM}").value = data
        
            # 이미지 셀 너비, 높이 설정
            self.ws_summary.range(f'B{self.ROW_NUM}').row_height = self.IMG_HEIGHTSIZE
            self.ws_summary.range(f'B{self.ROW_NUM}').column_width = self.IMG_FAINAL_WIDTH
            
            self.ROW_NUM += 1

    def set_summaryStyle(self, START_ROW):
        """SUMMARY시트에서 해당 언어 영역 스타일(테두리, 자동줄바꿈, 가운데 맞춤) 설정

        Args:
            START_ROW (_type_): 해당 언어가 시작되는 행
        """
        summaryTableRange = self.ws_summary.range(f"A{START_ROW}").expand('table')
        self.set_style(summaryTableRange)

    def set_langSheetStyle(self):
        """언어별 시트에서 스타일(테두리, 자동줄바꿈, 가운데 맞춤) 설정
        """
        tableRange = self.ws2.range("A1").expand('table')
        self.set_style(tableRange)

    def set_style(self, range):
        """테두리, 자동줄바꿈, 가운데 맞춤
        """
        self.alignCenter(range)
        self.setBorder(range)
        range.api.WrapText = True

    def excel_setup(self):
        """이미지 사이즈, 열너비를 정의하는 함수
        """
        excel_setList, _ = self.sp.read_setup(table = "Excel_Setting")

        self.IMG_WIDTHSIZE = int(excel_setList[0]) * 15 / 0.53   # 이미지 넓이
        self.IMG_HEIGHTSIZE = int(excel_setList[1]) * 15 / 0.53  # 이미지 높이
        self.SHEET_WIDTHSIZE = int(excel_setList[2])             # 필드 넓이
        self.SHEET_EvaluationListSIZE = int(excel_setList[3])    # 평가 목록 넓이
        self.IMG_FAINAL_WIDTH = self.IMG_WIDTHSIZE * 70.25 / 425
        self.IMG_SHEET_HEIGHTSIZE = self.IMG_HEIGHTSIZE
        
        # width = 0
        # if self.IMG_WIDTHSIZE <= 380:
        #     width = 400 - self. IMG_WIDTHSIZE
        #     width = width // 100 + 0.8
        # # 310 ~ 569
        # elif self.IMG_WIDTHSIZE >= 420:
        #     width = self. IMG_WIDTHSIZE - 400 
        #     width = width // 120 - 0.9
        #     width = - width

        # self.IMG_FAINAL_WIDTH = self.IMG_WIDTHSIZE // 8 + width # 하기와 같이 수정 필요 제한 크기사항도 필요
        # self.IMG_SHEET_HEIGHTSIZE = self.IMG_HEIGHTSIZE // 5 * 4 # 하기와 같이 수정 필요 제한 크기사항도 필요

        # if int(excel_setList[1]) == 155:
        #     self.IMG_HEIGHTSIZE = 155
        #     self.IMG_SHEET_HEIGHTSIZE = 115
            
    def create_sheet_history(self, lang:str, ws:object):
        """
        연속적으로 생성 할 수 있는 규칙 필요
        """

        lang_books = []
        lang_books.append("언어")

        
        cell_idx = 1
        cell_row = self.history_rows
        sequence = 0
        createBool = False
        for val in self.create_tupleBooks(lang=lang):
            lang_books.append(val)

        for idx, path in enumerate(self.create_imgCellCount(lang = lang)):
            if (path != ""):
                
                res = [ele for ele in self.fail_checkList if(ele in self.overloading_cell_search_val(sequence = idx, lang= lang)[:len(self.evaluation_len("Test_List")) + 1])]
                resval = self.overloading_cell_search_val(sequence = idx, lang= lang)[:len(self.evaluation_len("Test_List")) + 1]
                print(f"res create : {resval} {bool(res)}")
                if bool(res):
                    createBool = True
        
        if createBool:
            # 컬럼 설정
            for val in range(0, len(lang_books)):
                ws.cell(row = self.history_rows, # 현재 진행상황
                            column= cell_idx, # 평가한 나라 개수 영향
                            value=lang_books[val])
                cell_idx = cell_idx + 1
            self.history_rows = self.history_rows + 1
        else:
            return

        for idx, path in enumerate(self.create_imgCellCount(lang = lang)):

            if (path != ""):
                
                res = [ele for ele in self.fail_checkList if(ele in self.overloading_cell_search_val(sequence = idx, lang= lang)[:len(self.evaluation_len("Test_List")) + 1])]
                resval = self.overloading_cell_search_val(sequence = idx, lang= lang)[:len(self.evaluation_len("Test_List")) + 1]
                print(f"res create : {resval} {bool(res)}")
                if bool(res):

                    columns= 0 + 1
                    print(f'columns : {columns}')
                    print(f'Value : {lang}')
                    ws.cell(row = self.history_rows,
                            column= columns,
                            value = lang)

                    columns = columns + 1
                    columnsVal = 0
                    print(f'columns : {columns}')
                    ws.cell(row = self.history_rows,
                            column= columns,
                            value = self.cell_search_val(sequence = sequence, columns = columnsVal, lang = lang))
                    
                    # 평가 결과
                    for idx in range(0, len(self.evaluation_len(key="Test_List"))):

                        columns= columns + 1
                        columnsVal = columnsVal + 1
                        print(f'columns : {columns}')
                        print(f'Value : {self.cell_search_val(sequence = sequence, columns = columnsVal, lang = lang)}')
                        value = self.cell_search_val(sequence = sequence, columns = columnsVal, lang = lang)

                        ws.cell(row = self.history_rows,
                            column= columns,
                            value = ("" if "" == value else value))

                    # 라벨 평가
                    for idx in range(0, len(self.evaluation_len(key="Field"))):

                        columns= columns + 1
                        columnsVal = columnsVal + 1
                        print(f'columns : {columns}')
                        print(f'Value : {self.cell_search_val(sequence = sequence, columns = columnsVal, lang = lang)}')
                        value = self.cell_search_val(sequence = sequence, columns = columnsVal, lang = lang)

                        ws.cell(row = self.history_rows,
                            column= columns,
                            value = ("" if "" == value else value))

                
                    # 버전 정보
                    print(f'columns : {columns + 1}')
                    print(f'Value : {self.cell_search_val(sequence = sequence, columns = columnsVal + 1, lang = lang)}')
                    value = self.cell_search_val(sequence = sequence, columns = columnsVal + 1, lang = lang)
                    ws.cell(row = self.history_rows,
                                column= columns + 1,
                                value = "" if "" == value else value)

                    self.history_rows = self.history_rows + 1
                sequence = sequence + 1

        self.set_cellStyle(active= ws, idx = cell_row, fix= "SUMMARY")
        self.history_rows = self.history_rows + 1
        
    def excel_data_input(self, active : Worksheet, lang):
        # 데이터 추가
        
        lang_books = []

        cell_idx = 1
        cell_rows = 1
        for val in self.create_tupleBooks(lang=lang):
            lang_books.append(val)

        # 태그적용
        for val in range(0, len(lang_books)):
            active.cell(row = cell_rows, # 현재 진행상황
                        column= cell_idx, # 평가한 나라 개수 영향
                        value=lang_books[val])
            cell_idx = cell_idx + 1

        self.set_cellStyle(active= active, idx = cell_rows, fix= None)

        # 언어 이미지명 testList Field 버전
        
        sequence = 0
        cell_rows = cell_rows + 1
        self.imgList = self.create_imgCellCount(lang = lang)
        for i, path in enumerate(self.imgList):

            if (path != ""):

                columns= 1
                columnsVal = 0
                
                try:
                    img = Image(path)
                    img.width = self.IMG_WIDTHSIZE
                    img.height = self.IMG_HEIGHTSIZE
                    
                    active.add_image(img=img, anchor=f"A{cell_rows}") # 이미지 추가
                    active.row_dimensions[cell_rows].height = self.IMG_SHEET_HEIGHTSIZE
                    active.cell(row = cell_rows,
                                column= columns,
                                value = self.cell_search_val(sequence = sequence, columns = columnsVal, lang = lang))
                except:
                    active.row_dimensions[cell_rows].height = self.IMG_SHEET_HEIGHTSIZE
                    active.cell(row = cell_rows,
                                column= idx + 1,
                                value = "파일 없음")
                active.cell(row = cell_rows,
                        column= columns,
                        value = self.cell_search_val(sequence = sequence, columns = columnsVal, lang = lang))
                
                # 평가 결과
                for idx in range(0, len(self.evaluation_len(key="Test_List"))):

                    columns= columns + 1
                    columnsVal = columnsVal + 1
                    print(f'columns : {columns}')
                    print(f'Value : {self.cell_search_val(sequence = sequence, columns = columnsVal, lang = lang)}')
                    value = self.cell_search_val(sequence = sequence, columns = columnsVal, lang = lang)

                    active.cell(row = cell_rows,
                        column= columns,
                        value = ("" if "" == value else value))

                # 라벨 평가
                for idx in range(0, len(self.evaluation_len(key="Field"))):

                    columns= columns + 1
                    columnsVal = columnsVal + 1
                    print(f'columns : {columns}')
                    print(f'Value : {self.cell_search_val(sequence = sequence, columns = columnsVal, lang = lang)}')
                    value = self.cell_search_val(sequence = sequence, columns = columnsVal, lang = lang)

                    active.cell(row = cell_rows,
                        column= columns,
                        value = ("" if "" == value else value))

            
                # 버전 정보
                print(f'columns : {columns + 1}')
                print(f'Value : {self.cell_search_val(sequence = sequence, columns = columnsVal + 1, lang = lang)}')
                value = self.cell_search_val(sequence = sequence, columns = columnsVal + 1, lang = lang)
                active.cell(row = cell_rows,
                            column= columns + 1,
                            value = "" if "" == value else value)

                cell_rows = cell_rows + 1
            
            self.set_cellStyle(active= active, idx= 1, fix=None)
            sequence = sequence + 1
            percent_val = round((self.start_percent + ((i+1)/len(self.imgList))*self.split_percent)*100)
            if percent_val == 100:
                percent_val = 99
            self.progressBarValue.emit(percent_val)
            
        cell_rows = cell_rows + 1
        
    def set_cellStyle(self, active : Worksheet, idx : int, fix:str):

        i = 0
        for column in active.columns:
            
            if fix == "SUMMARY":
                if (i == 1 or i > len(self.evaluation_len(key="Test_List")) + 1):
                    active.column_dimensions[self.column[i]].width = self.SHEET_WIDTHSIZE
                else:
                    active.column_dimensions[self.column[i]].width = self.SHEET_WIDTH_SHORTSIZE
            else:
                if i == 0:
                    active.column_dimensions[self.column[0]].width = self.IMG_FAINAL_WIDTH
                elif (i > len(self.evaluation_len(key="Test_List"))):
                    active.column_dimensions[self.column[i]].width = self.SHEET_WIDTHSIZE
                else:
                    active.column_dimensions[self.column[i]].width = self.SHEET_WIDTH_SHORTSIZE

            active[f'{self.column[i]}{idx}'].font = Font(size=11)

            i = i + 1

        upper = [f'{i}{idx}' for i in string.ascii_uppercase]
        
        for val_row in active.rows:
            for cell in val_row:

                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                str_cell = str(cell)
                str_cell = str_cell[str_cell.find(".") + 1:str_cell.find(">")]
                print(f'cell : {str_cell}')
                print(f'upper : {upper}')
                if (str_cell in upper):
                    cell.fill = PatternFill(patternType="solid", fgColor="99CC00")

                cell.border = Border(
                            right=Side(border_style=borders.BORDER_THIN,
                                    color='000000'),
                            bottom=Side(border_style=borders.BORDER_THIN,
                                        color='000000'))
                time.sleep(0.005)
                QApplication.processEvents()
        print(f'upperEXIT')
        
    def create_imgCellCount(self, lang) -> List:
        """
        1. DB 받아오기
        2. 저장된 DB중 이미지 경로 있는것만 구별시키기
        3. 이미지 개수를 확인하고 : return 해당경로를 순서대로 List 반환
        """

        # self.c.execute(f"SELECT * FROM '{lang}'")
        # dataList = self.c.fetchall()
        dataList = db.db_select(f"SELECT * FROM '{lang}'")
    
        idx = 0 # "평가결과저장된데이터중 경로위치"
        img_pathList = []

        # 행마다 차래로
        for data in dataList:
            
            if data != "":
                # clume서치
                img_pathList.append(data[idx])

        return img_pathList
    
    def create_tupleBooks(self, lang) -> list:
        """
        언어 마다 엑셀 타이틀 항목 반환 함수
        * ini에 영향 끼침.

        return : Title list result
        """

        tupleBooks = {}

        for key in self.lang_List:
            
            all_List = []
            all_List.append(f'{key}_이미지')

            for val in self.evaluation_len(key="Test_List"):
                all_List.append(val)

            for val in self.evaluation_len(key="Field"):
                all_List.append(val)
            
            all_List.append("버전 정보")

            tupleBooks.setdefault(key, all_List)
        
        return tupleBooks.get(lang)
    
    def evaluation_len(self, key) -> List:
    
        result_val, result_val2 = self.sp.read_setup(table = key)
  
        return result_val
    
    def overloading_cell_search_val(self, sequence, lang) -> List:
        """
        1. DB 받아오기
        2. 저장된 DB중 이미지 경로 있는것만 구별시키기
        3. 이미지 개수를 확인하고 : return 해당경로를 순서대로 List 반환
        """

        # self.c.execute(f"SELECT * FROM '{lang}'")
        # dataList = self.c.fetchall()
        dataList = db.db_select(f"SELECT * FROM '{lang}'")

        return dataList[sequence]
    
    def cell_search_val(self, sequence, columns, lang) -> List:
        """
        1. DB 받아오기
        2. 저장된 DB중 이미지 경로 있는것만 구별시키기
        3. 이미지 개수를 확인하고 : return 해당경로를 순서대로 List 반환
        """

        # self.c.execute(f"SELECT * FROM '{lang}'")
        # dataList = self.c.fetchall()
        dataList = db.db_select(f"SELECT * FROM '{lang}'")
        print(f'{dataList[sequence]} : {dataList[sequence][columns]}')

        return dataList[sequence][columns]
        

    def __equalsVerification__(self, kr_list : list, ws : object, lang : str, terget : int):
        """
        구별 : 차별필요
        """
        lang_books = self.create_tupleBooks(lang=lang)

        if len(kr_list) > len(lang_books):

            del_ColCount = len(kr_list) - (len(kr_list) - len(lang_books)) + 1
            for i in range(0, len(kr_list) - len(lang_books)):
                del kr_list[len(kr_list) - 1]
                ws.Columns(del_ColCount).EntireColumn.Delete()
        elif len(kr_list) < len(lang_books):
            for i in range(0, len(lang_books) - len(kr_list)):
                kr_list.append("NULL")

        # 태그적용
        for val in range(0, len(kr_list)):
            cell_idx = val + 1
            if kr_list[val] != lang_books[val]:
                ws.Cells(terget, cell_idx).Value = lang_books[val]
                kr_list[val] = lang_books[val]
            else:
                ws.Cells(terget, cell_idx).Value = kr_list[val]
            ws.Cells(terget, cell_idx).Interior.ColorIndex = self.TABLE_CELL_COLOR # 색상적용

            print(f"StyleCell1 : {terget}, {cell_idx}")
            self.set_Win32com_cellStyle(ws = ws, terget = terget, row_idx = cell_idx, heightSize = None, columnsList=kr_list)

        getPathList = os.path.dirname(self.create_imgCellCount(lang)[0])
        cnt = 0
        fileEx = r'.png'
        xlsx_list = [os.path.join(getPathList, file) for file in os.listdir(getPathList) if file.endswith(fileEx)]

        # 엑셀 데이터 적용
        for key_terget in range(0, len(self.create_imgCellCount(lang=lang))):
            
            time.sleep(0.003)
            cellCount = key_terget + 2
            maxCount = len(self.create_imgCellCount(lang=lang)) + 1
            if cellCount <= maxCount:

                path = self.create_imgCellCount(lang=lang)[key_terget]
                if os.path.isfile(path= path):
                    getStr = str(xlsx_list[key_terget])
                    getPath = getStr.replace('/', '\\', 10)
                    rng = ws.Range(f"A{cellCount}")
                    shp = ws.Shapes.AddPicture(
                        Filename=rf"{getPath}",
                        LinkToFile=False,
                        SaveWithDocument=True,
                        Left=rng.Left,
                        Top=rng.Top,
                        Width= (self.IMG_WIDTHSIZE - (self.IMG_FAINAL_WIDTH * 2)),
                        Height=self.IMG_SHEET_HEIGHTSIZE
                    )
                else:
                    print("없음")
                ws.Cells(cellCount, 1).Value = path

                data_List = list(self.overloading_cell_search_val(sequence= key_terget, lang= lang))

                print(f'kr_list확인 : {kr_list}')
                for val in range(0, len(kr_list)):
                    cell_idx = val + 1
                    ws.Cells(cellCount, cell_idx).Value = (" " if "" == data_List[val] or " " == data_List[val] else data_List[val])
                    print(f"wsValue : {ws.Cells(cellCount, cell_idx).Value}")
                    print(f"StyleCell2 : {cellCount}, {cell_idx}")
                    self.set_Win32com_cellStyle(ws = ws, terget = cellCount, row_idx = cell_idx,\
                         heightSize = self.IMG_SHEET_HEIGHTSIZE, columnsList=kr_list)
                    time.sleep(0.003)
                print(f'진행도 : {cellCount}/{maxCount}')
                QApplication.processEvents()
            percent_val = round((self.start_percent + ((key_terget+1)/len(self.create_imgCellCount(lang=lang)))*self.split_percent)*100)
            print(f'percent_val : {percent_val}')
            if percent_val == 100:
                percent_val = 99
            self.progressBarValue.emit(percent_val)
            print(f'percent_val : {percent_val}####')

    def set_Win32com_cellStyle(self, ws : object, terget : int, row_idx : int, heightSize : int, columnsList : list):

        print(f'self.IMG_FAINAL_WIDTH : {self.IMG_FAINAL_WIDTH}')
        print(f'self.IMG_SHEET_HEIGHTSIZE : {heightSize}')
        print(f'self.SHEET_WIDTHSIZE : {self.SHEET_WIDTHSIZE}')
        print(f'self.SHEET_EvaluationListSIZE : {self.SHEET_EvaluationListSIZE}')

        upper = [f'{i}1' for i in string.ascii_uppercase]
        if ws.Name == "SUMMARY":
            len_TestList = len(self.evaluation_len(key="Test_List")) + 2
            wbNumber = 20
        else:
            len_TestList = len(self.evaluation_len(key="Test_List")) + 1
            wbNumber = self.IMG_FAINAL_WIDTH

        if heightSize != None:

            if row_idx > 1:
                ws.Rows(terget).RowHeight = heightSize  # 선택 영역 행 크기 설정

            if row_idx < len_TestList:
                
                result = ws.Range(f"{upper[row_idx]}").Value
                if isinstance(result, int) or isinstance(result, float):
                    result = str(result)

                if len(result) > 10:
                    ws.Range(f"{upper[row_idx]}").ColumnWidth = self.SHEET_WIDTHSIZE
                else:
                    ws.Range(f"{upper[row_idx]}").ColumnWidth = self.SHEET_EvaluationListSIZE
            elif len_TestList <= row_idx and row_idx <= len(columnsList):
                ws.Range(f"{upper[row_idx]}").ColumnWidth = self.SHEET_WIDTHSIZE

        ws.Cells(terget, 1).ColumnWidth = wbNumber
        ws.Cells(terget, row_idx).Font.Size = 11
        ws.Cells(terget, row_idx).VerticalAlignment = -4108  #가운데 정렬(수직)
        ws.Cells(terget, row_idx).HorizontalAlignment = -4108  #가운데 정렬(수평)
        
        rng = ws.Cells(terget, row_idx) # 사용 영역 선택
    
        #테두리 설정
        rng.Borders.LineStyle = 1 #선 스타일
        rng.Borders.ColorIndex = 1 #선 색상 : Black
        rng.Borders.Weight = 2 #선 굵기

        try:
            if ws.Cells(terget, 1).Value == "언어":
                ws.Rows(terget).RowHeight = 20  # 선택 영역 행 크기 설정
                if row_idx > len(columnsList):
                    print(f'terget : {terget}, {upper[row_idx]}')
                    ws.Range(f"{upper[len(columnsList)]}").ColumnWidth = 8.38
            else:
                ws.Range(f"{upper[len(columnsList)]}").ColumnWidth = 8.38
            
        except:
            print("summmary update영역아님")

    def update_sheet_history(self, lang:str, del_lastCell:bool, integer:int):

        def passing_ex():
            # 중복방지
            count = 1
            self.historyUpdate_rows = self.historyUpdate_rows + 1
            while(True):
                print(f"확인 historyUpdate_rows : {self.historyUpdate_rows}")
                print(f"확인 : {ws.Cells(self.historyUpdate_rows, integer).Value}")
                if count == 10:
                    print("카운트 멈춤")
                    break
                elif ws.Cells(self.historyUpdate_rows, integer).Value == "언어" and ws.Cells(self.historyUpdate_rows + 1, integer).Value == lang:
                    print(f"삭제1 : {ws.Cells(self.historyUpdate_rows, integer).Value}")
                    ws.Rows(self.historyUpdate_rows).Delete()
                    self.historyUpdate_rows = self.historyUpdate_rows - 1
                    count = 0
                elif ws.Cells(self.historyUpdate_rows, integer).Value == lang:
                    print(f"삭제2 : {ws.Cells(self.historyUpdate_rows, integer).Value}")
                    ws.Rows(self.historyUpdate_rows).Delete()
                    self.historyUpdate_rows = self.historyUpdate_rows - 1
                    count = 0
                elif ws.Cells(self.historyUpdate_rows, integer).Value == None:
                    print(f"카운트 적립 : {count}")
                    count = count + 1
                
                # 빈 행 맞추기 위한 조치에 의한 라인 중복 삭제
                if ws.Cells(self.historyUpdate_rows - 1, integer).Value == None and ws.Cells(self.historyUpdate_rows, integer).Value == None:
                    print(f"삭제3 : {ws.Cells(self.historyUpdate_rows, integer).Value}")
                    ws.Rows(self.historyUpdate_rows).Delete()
                    self.historyUpdate_rows = self.historyUpdate_rows - 1
                time.sleep(0.003)

                self.historyUpdate_rows = self.historyUpdate_rows + 1

            if del_lastCell:
                ws.Rows(self.historyUpdate_rows - 1).Delete()

        ws = self.wb.Worksheets("SUMMARY")
        row = self.historyUpdate_rows
        colCount = 1
        colList = []
        getPathList = []
        pathList = []
        tuples = []
        tuples.append("언어")

        for tuple in self.create_tupleBooks(lang= lang):
            tuples.append(tuple)

        # 새로운 추가 건수
        for idx in range(0, len(self.create_imgCellCount(lang = lang))):
            data_List = list(self.overloading_cell_search_val(sequence= idx, lang= lang))

            res = [ele for ele in self.fail_checkList if(ele in data_List[:len(self.evaluation_len("Test_List")) + 2])]
            resval = data_List[:len(self.evaluation_len("Test_List")) + 2]
            print(f"res : {resval} {bool(res)}")
            if bool(res):
                pathList.append(os.path.basename(self.create_imgCellCount(lang)[idx])) # 파일명만 추출모음

        # 빈셀 삭제 작업
        if len(pathList) == 0:

            key = integer
            while(ws.Cells(integer, integer).Value == None):
                ws.Rows(integer).Delete()
                if key == 10:
                    break
                else:
                    key = key + 1
            print("passingex")
            passing_ex()
            return
        else:
            key = integer
            while(ws.Cells(integer, integer).Value == None):
                ws.Rows(integer).Delete()
                if key == 10:
                    break
                else:
                    key = key + 1

        # 빈 행 맞추기 위한 조치
        for i in range(2):
            ws.Rows(1).EntireRow.Insert()

        while(ws.Cells(row, integer).Value == lang):
            print(f"확인 : {ws.Cells(row, integer).Value}")

            row = row + 1
        
        # 비교파일명만 저장
        for row_idx in range(self.historyUpdate_rows, row):
            print(f"ws.Cells(row_idx, integer).Value : {ws.Cells(row_idx, integer).Value}")
            if (ws.Cells(row_idx, integer).Value == lang):
                getPathList.append(os.path.basename(ws.Cells(row_idx, 2).Value)) 

        # 각 차이좀 조사 및 행 추가 삭제
        if len(getPathList) != len(pathList):
            print(f"ERROR|pathList : {len(pathList)}")
            char_lenght = len(pathList) - len(getPathList)
            if (char_lenght > 0):
                
                for i in range(0, char_lenght):
                    ws.Rows(row).Insert()
                    
            elif (char_lenght < 0):
                for i in range(0, abs(char_lenght)):
                    ws.Rows(row).Delete()

        row = row - 1
        self.historyUpdate_rows = self.historyUpdate_rows - 1

        for i in range(0, row):
            ws.Rows(self.historyUpdate_rows).Delete()
        
        for i in range(0, row):
            ws.Rows(self.historyUpdate_rows).Insert()
        
        # 태그적용
        for val in range(0, len(tuples)):
            print(f"tuples[val] : {tuples[val]}")
            print(f"historyUpdate_rows : {self.historyUpdate_rows}")
            ws.Cells(self.historyUpdate_rows, val + 1).Value = tuples[val]
            ws.Cells(self.historyUpdate_rows, val + 1).Interior.ColorIndex = self.TABLE_CELL_COLOR # 색상적용
            self.set_Win32com_cellStyle(ws = ws, terget = self.historyUpdate_rows, row_idx = val + 1\
                , heightSize = None, columnsList=tuples)
            
        # 엑셀 데이터 적용 NA NT FAIL 분류하기
        for key_terget in range(0, len(self.create_imgCellCount(lang=lang))):

            data_List = list(self.overloading_cell_search_val(sequence= key_terget, lang= lang))
            res = [ele for ele in self.fail_checkList if(ele in data_List[:len(self.evaluation_len("Test_List")) + 2])]
            resval = data_List[:len(self.evaluation_len("Test_List")) + 2]
            print(f"res : {resval} {bool(res)}")
            if bool(res):
                self.historyUpdate_rows = self.historyUpdate_rows + 1
                resultList = []
                resultList.append(lang)
                for key in data_List:
                    resultList.append(key)

                print(f"resultList : {resultList}")
                for val in range(0, len(resultList)):
                    ws.Cells(self.historyUpdate_rows, val + 1).Value = (" " if "" == resultList[val] or " " == resultList[val] else resultList[val])
                    self.set_Win32com_cellStyle(ws = ws, terget = self.historyUpdate_rows, row_idx = val + 1\
                        , heightSize = self.IMG_SHEET_HEIGHTSIZE, columnsList=resultList)
                
            print(f'RESULT 진행도 : {key_terget}/{len(self.create_imgCellCount(lang=lang))}')
            time.sleep(0.005)
            QApplication.processEvents()

        passing_ex()

# class excelRun(QWidget):

#     def __init__(self, save_path, lang_List, new_set_difference) -> None:
#         super().__init__()

#         wb = object
#         COUNT = 0
#         ADD_COUNT = 0
#         self.save_path = save_path
#         print(save_path)
#         for lang in lang_List:
#             COUNT = COUNT + len(self.imgCellCount(lang= lang))

#         if new_set_difference:
#             wb = xl.Workbook()
#             ADD_COUNT = 50
#         else:
#             ADD_COUNT = 70
#             COUNT = COUNT + ADD_COUNT

#         if COUNT > 100:

#             for i in range(COUNT // 100):
#                 COUNT = COUNT + ADD_COUNT
        
#         if (self.setting_Verification(langList = lang_List)):
#             self.progress_Thread = QThread()
#             self.progress_Thread.start()
#             self.worker = ProgressApp(time=int(COUNT), new_set_difference = new_set_difference, save_path= self.save_path, wb= wb)
#             self.worker.moveToThread(self.progress_Thread)

#             self.exModul_Thread = QThread()
#             self.exModul_Thread.start()
#             self.exModuls = excelModul(save_path = self.save_path, lang_List = lang_List, new_set_difference = new_set_difference, wb= wb)
#             self.exModuls.moveToThread(self.exModul_Thread)

#     def setting_Verification(self, langList):
#         path = str(os.path.dirname(self.save_path))

#         result = False
        
#         if os.path.isdir(path):
#             result = True
#         else:
#             btnReply = QMessageBox.warning(self, "주의", f"{path} 경로가 존재하지 않습니다.", QMessageBox.Ok, QMessageBox.Ok)
#             LogManager.HLOG.info("엑셀 생성 팝업에서 존재하지 않는 경로 알림 표시")
            
#             if btnReply == QMessageBox.Ok:
#                 result = False
#                 return result

#         for lang in langList:
                
#             path = os.path.dirname(self.imgCellCount(lang))
#             if os.path.isdir(path):
#                 result = True
#             else:
#                 btnReply = QMessageBox.warning(self, "주의", f"{self.imgCellCount(lang)[0]} 경로가 존재하지 않습니다.", QMessageBox.Ok, QMessageBox.Ok)
#                 LogManager.HLOG.info("언어 설정 팝업에서 존재하지 않는 경로 알림 표시")
#                 if btnReply == QMessageBox.Ok:
#                     result = False
#                     return result
    
#         return result

#     def imgCellCount(self, lang) -> List:
#         """
#         1. DB 받아오기
#         2. 저장된 DB중 이미지 경로 있는것만 구별시키기
#         3. 이미지 개수를 확인하고 : return 해당경로를 순서대로 List 반환
#         """

#         # self.c.execute(f"SELECT * FROM '{lang}'")
#         # dataList = self.c.fetchall()
#         data = db.db_select_one(f"SELECT * FROM '{lang}'")
    
#         # idx = 0 # "평가결과저장된데이터중 경로위치"
#         # img_pathList = []

#         # # 행마다 차래로
#         # for data in dataList:
            
#         #     if data != "":
#         #         # clume서치
#         #         img_pathList.append(data[idx])

#         return data[0]

# class excelModul():
    
#     # def __init__(self, save_path, lang_List, new_set_difference, wb:object) -> excel:
#     def __init__(self, parent=None):
#         # super().__init__(parent)
#         self.save_path = parent.path
#         self.lang_List = parent.selected_langList
#         self.new_set_difference = parent.testBool
#         self.wb = parent.wb

#         # loop = QEventLoop()
#         # QTimer.singleShot(1000, loop.quit)

#         # 초기설정
#         self.IMG_WIDTHSIZE = 400 # 이미지 넓이
#         self.IMG_HEIGHTSIZE = 150 # 이미지 높이
#         self.SHEET_WIDTHSIZE = None # 필드 넓이
#         self.SHEET_HEIGHTSIZE = None # 필드 높이
#         self.IMG_SHEET_HEIGHTSIZE = 115 # 이미지 간격
#         self.IMG_FAINAL_WIDTH = 50
#         self.SHEET_WIDTH_SHORTSIZE = 15 # 시트 열 기본 작은 크기
#         self.TABLE_CELL_COLOR = 43 # 테이블 컬러
        
#         self.set_row = 1
#         self.start_column = 1
#         self.lang_cnt = 1
#         self.sp = sp.Settings()
#         self.get_lang_List = self.lang_List

#         self.column = []

#         self.excel_setup()
#         print(f'self.IMG_WIDTHSIZE : {self.IMG_WIDTHSIZE}')
#         print(f'self.IMG_HEIGHTSIZE : {self.IMG_HEIGHTSIZE}')
#         print(f'self.IMG_SHEET_HEIGHTSIZE : {self.IMG_SHEET_HEIGHTSIZE}')
#         print(f'self.SHEET_WIDTHSIZE : {self.SHEET_WIDTHSIZE}')
#         print(f'self.SHEET_HEIGHTSIZE : {self.SHEET_HEIGHTSIZE}')

#         for i in range(65, 91):
#             self.column.append(chr(i))

#         # self.run_excel()

#         # if (self.new_set_difference):

#         #     self.create_excel()
#         #     self.save_excel()

#         # else:
#         #     pass

#             # TITLE_TERGET = 1
#             # self.historyUpdate_rows = 2

#             # excel = win32.Dispatch("Excel.Application")
#             # excel.Visible = False
#             # self.wb = excel.Workbooks.Open(save_path)

#             # for lang in lang_List:
#             #     self.update_sheet_history(lang, self.wb.Worksheets("SUMMARY"), TITLE_TERGET)
#             #     ws = self.wb.Worksheets(lang)

#             #     kr_list = [] # 현재 엑셀 데이터
#             #     idx = 1

#             #     while(ws.Cells(TITLE_TERGET, idx).Value != None):
#             #         value = "" if "" == ws.Cells(TITLE_TERGET, idx).Value else ws.Cells(TITLE_TERGET, idx).Value
                    
#             #         kr_list.append(value)
#             #         idx = idx + 1

#             #     self.__equalsVerification__(kr_list=kr_list, ws = self.wb.Worksheets(lang), lang = lang, terget=TITLE_TERGET)
#             #     QApplication.processEvents()
                
#             # excel.Quit()

#         # loop.exec_()
#         # print(f'루프종료')
#         # loop.exit()
#         # signal_done.emit(1)

#     def create_excel(self):
#         ws = self.wb.active
#         ws.title = "SUMMARY"
#         self.history_rows = 1

#         for i in range(101):
#             yield i

#         # for idx, lang in enumerate(self.lang_List):

#         #     self.create_sheet_history(lang=lang, ws= ws)

#         #     active = self.wb.create_sheet(title=lang)

#         #     self.excel_data_input(active = active, lang = lang)
            
#         #     self.set_cellStyle(active= active, idx= 1, fix=None)

#         #     yield (idx+1)/(len(self.lang_List)+1)*100
            
#             # progressBarValue.emit((idx+1)/(len(lang_List)+1)*100)

#     def save_excel(self):
#         self.wb.save(f'{self.save_path}\\excelTest.xlsx')
#         return 1

#     def excel_setup(self):
#         excel_setList, _ = self.sp.read_setup(table = "Excel_Setting")

#         self.IMG_WIDTHSIZE = (int(excel_setList[0]) // 10) * 10 # 이미지 넓이
#         self.IMG_HEIGHTSIZE = (int(excel_setList[1]) // 10) * 10 # 이미지 높이
#         self.SHEET_WIDTHSIZE = int(excel_setList[2]) # 필드 넓이
#         self.SHEET_HEIGHTSIZE = int(excel_setList[3]) # 평가 목록 넓이
        
#         width = 0
#         if self.IMG_WIDTHSIZE <= 380:
#             width = 400 - self. IMG_WIDTHSIZE
#             width = width // 100 + 0.8
#         # 310 ~ 569
#         elif self.IMG_WIDTHSIZE >= 420:
#             width = self. IMG_WIDTHSIZE - 400 
#             width = width // 120 - 0.9
#             width = - width

#         self.IMG_FAINAL_WIDTH = self.IMG_WIDTHSIZE // 8 + width # 하기와 같이 수정 필요 제한 크기사항도 필요
#         self.IMG_SHEET_HEIGHTSIZE = self.IMG_HEIGHTSIZE // 5 * 4 # 하기와 같이 수정 필요 제한 크기사항도 필요

#         if int(excel_setList[1]) == 155:
#             self.IMG_HEIGHTSIZE = 155
#             self.IMG_SHEET_HEIGHTSIZE = 115


#     def create_imgCellCount(self, lang) -> List:
#         """
#         1. DB 받아오기
#         2. 저장된 DB중 이미지 경로 있는것만 구별시키기
#         3. 이미지 개수를 확인하고 : return 해당경로를 순서대로 List 반환
#         """

#         # self.c.execute(f"SELECT * FROM '{lang}'")
#         # dataList = self.c.fetchall()
#         dataList = db.db_select(f"SELECT * FROM '{lang}'")
    
#         idx = 0 # "평가결과저장된데이터중 경로위치"
#         img_pathList = []

#         # 행마다 차래로
#         for data in dataList:
            
#             if data != "":
#                 # clume서치
#                 img_pathList.append(data[idx])

#         return img_pathList

#     def cell_search_val(self, sequence, columns, lang) -> List:
#         """
#         1. DB 받아오기
#         2. 저장된 DB중 이미지 경로 있는것만 구별시키기
#         3. 이미지 개수를 확인하고 : return 해당경로를 순서대로 List 반환
#         """

#         # self.c.execute(f"SELECT * FROM '{lang}'")
#         # dataList = self.c.fetchall()
#         dataList = db.db_select(f"SELECT * FROM '{lang}'")
#         print(f'{dataList[sequence]} : {dataList[sequence][columns]}')

#         return dataList[sequence][columns]

#     def overloading_cell_search_val(self, sequence, lang) -> List:
#         """
#         1. DB 받아오기
#         2. 저장된 DB중 이미지 경로 있는것만 구별시키기
#         3. 이미지 개수를 확인하고 : return 해당경로를 순서대로 List 반환
#         """

#         # self.c.execute(f"SELECT * FROM '{lang}'")
#         # dataList = self.c.fetchall()
#         dataList = db.db_select(f"SELECT * FROM '{lang}'")

#         return dataList[sequence]

#     def evaluation_len(self, key) -> List:

#         result_val, result_val2 = self.sp.read_setup(table = key)
        
#         result_cellList = []

#         for i in result_val:
#             result_cellList.append(i)
  
#         return result_cellList

#     def excel_data_input(self, active : Worksheet, lang):
#         # 데이터 추가
        
#         lang_books = []

#         cell_idx = 1
#         cell_rows = 1
#         for val in self.create_tupleBooks(lang=lang):
#             lang_books.append(val)

#         # 태그적용
#         for val in range(0, len(lang_books)):
#             active.cell(row = cell_rows, # 현재 진행상황
#                         column= cell_idx, # 평가한 나라 개수 영향
#                         value=lang_books[val])
#             cell_idx = cell_idx + 1

#         self.set_cellStyle(active= active, idx = cell_rows, fix= None)

#         # 언어 이미지명 testList Field 버전
        
#         sequence = 0
#         cell_rows = cell_rows + 1
#         for path in self.create_imgCellCount(lang = lang):

#             if (path != ""):

#                 columns= 1
#                 columnsVal = 0
                
#                 try:
#                     img = Image(path)
#                     img.width = self.IMG_WIDTHSIZE
#                     img.height = self.IMG_HEIGHTSIZE
                    
#                     active.add_image(img=img, anchor=f"A{cell_rows}") # 이미지 추가
#                     active.row_dimensions[cell_rows].height = self.IMG_SHEET_HEIGHTSIZE
#                     active.cell(row = cell_rows,
#                                 column= columns,
#                                 value = self.cell_search_val(sequence = sequence, columns = columnsVal, lang = lang))
#                 except:
#                     active.row_dimensions[cell_rows].height = self.IMG_SHEET_HEIGHTSIZE
#                     active.cell(row = cell_rows,
#                                 column= idx + 1,
#                                 value = "경로 없음")
#                 active.cell(row = cell_rows,
#                         column= columns,
#                         value = self.cell_search_val(sequence = sequence, columns = columnsVal, lang = lang))
                
#                 # 평가 결과
#                 for idx in range(0, len(self.evaluation_len(key="Test_List"))):

#                     columns= columns + 1
#                     columnsVal = columnsVal + 1
#                     print(f'columns : {columns}')
#                     print(f'Value : {self.cell_search_val(sequence = sequence, columns = columnsVal, lang = lang)}')
#                     value = self.cell_search_val(sequence = sequence, columns = columnsVal, lang = lang)

#                     active.cell(row = cell_rows,
#                         column= columns,
#                         value = ("" if "" == value else value))

#                 # 라벨 평가
#                 for idx in range(0, len(self.evaluation_len(key="Field"))):

#                     columns= columns + 1
#                     columnsVal = columnsVal + 1
#                     print(f'columns : {columns}')
#                     print(f'Value : {self.cell_search_val(sequence = sequence, columns = columnsVal, lang = lang)}')
#                     value = self.cell_search_val(sequence = sequence, columns = columnsVal, lang = lang)

#                     active.cell(row = cell_rows,
#                         column= columns,
#                         value = ("" if "" == value else value))

            
#                 # 버전 정보
#                 print(f'columns : {columns + 1}')
#                 print(f'Value : {self.cell_search_val(sequence = sequence, columns = columnsVal + 1, lang = lang)}')
#                 value = self.cell_search_val(sequence = sequence, columns = columnsVal + 1, lang = lang)
#                 active.cell(row = cell_rows,
#                             column= columns + 1,
#                             value = "" if "" == value else value)

#                 cell_rows = cell_rows + 1
#             sequence = sequence + 1

#         cell_rows = cell_rows + 1


#     def set_cellStyle(self, active : Worksheet, idx : int, fix:str):

#         i = 0
#         for column in active.columns:
            
#             if fix == "SUMMARY":
#                 if (i == 1 or i > len(self.evaluation_len(key="Test_List")) + 1):
#                     active.column_dimensions[self.column[i]].width = self.SHEET_WIDTHSIZE
#                 else:
#                     active.column_dimensions[self.column[i]].width = self.SHEET_WIDTH_SHORTSIZE
#             else:
#                 if i == 0:
#                     active.column_dimensions[self.column[0]].width = self.IMG_FAINAL_WIDTH
#                 elif (i > len(self.evaluation_len(key="Test_List"))):
#                     active.column_dimensions[self.column[i]].width = self.SHEET_WIDTHSIZE
#                 else:
#                     active.column_dimensions[self.column[i]].width = self.SHEET_WIDTH_SHORTSIZE

#             active[f'{self.column[i]}{idx}'].font = Font(size=11)

#             i = i + 1

#         upper = [f'{i}{idx}' for i in string.ascii_uppercase]
        
#         for val_row in active.rows:
#             for cell in val_row:

#                 cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
#                 str_cell = str(cell)
#                 str_cell = str_cell[str_cell.find(".") + 1:str_cell.find(">")]
#                 print(f'cell : {str_cell}')
#                 print(f'upper : {upper}')
#                 if (str_cell in upper):
#                     cell.fill = PatternFill(patternType="solid", fgColor="99CC00")

#                 cell.border = Border(
#                             right=Side(border_style=borders.BORDER_THIN,
#                                     color='000000'),
#                             bottom=Side(border_style=borders.BORDER_THIN,
#                                         color='000000'))
#                 time.sleep(0.005)
#                 QApplication.processEvents()
#         print(f'upperEXIT')


#     def set_Win32com_cellStyle(self, ws : object, terget : int, cell_idx : int, heightSize : int):
#         if heightSize != None:

#             if cell_idx > 1:
#                 ws.Rows(cell_idx).RowHeight = heightSize  # 선택 영역 행 크기 설정
#         else:
#             ws.Rows(cell_idx).RowHeight = self.SHEET_HEIGHTSIZE  # 선택 영역 행 크기 설정

#         rng = ws.UsedRange # 사용 영역 선택
    
#         #테두리 설정
#         rng.Borders.LineStyle = 1 #선 스타일
#         rng.Borders.ColorIndex = 1 #선 색상 : Black
#         rng.Borders.Weight = 2 #선 굵기
#         ws.Cells(terget, cell_idx).Font.Size = 11
#         ws.Cells(terget, cell_idx).VerticalAlignment = -4108  #가운데 정렬(수직)
#         ws.Cells(terget, cell_idx).HorizontalAlignment = -4108  #가운데 정렬(수평)
    
#     def create_tupleBooks(self, lang) -> list:
#         """
#         언어 마다 엑셀 타이틀 항목 반환 함수
#         * ini에 영향 끼침.

#         return : Title list result
#         """

#         tupleBooks = {}

#         for key in self.get_lang_List:
            
#             all_List = []
#             all_List.append(f'{key}_이미지')

#             for val in self.evaluation_len(key="Test_List"):
#                 all_List.append(val)

#             for val in self.evaluation_len(key="Field"):
#                 all_List.append(val)
            
#             all_List.append("버전 정보")

#             tupleBooks.setdefault(key, all_List)
        
#         return tupleBooks.get(lang)

# def __equalsVerification__(self, kr_list : list, ws : object, lang : str, terget : int):
#     """
#     구별 : 차별필요
#     """
#     lang_books = self.create_tupleBooks(lang=lang)

#     if len(kr_list) > len(lang_books):

#         del_ColCount = len(kr_list) - (len(kr_list) - len(lang_books)) + 1
#         for i in range(0, len(kr_list) - len(lang_books)):
#             del kr_list[len(kr_list) - 1]
#             ws.Columns(del_ColCount).EntireColumn.Delete()
#     elif len(kr_list) < len(lang_books):
#         for i in range(0, len(lang_books) - len(kr_list)):
#             kr_list.append("NULL")

#     # 태그적용
#     for val in range(0, len(kr_list)):
#         cell_idx = val + 1
#         if kr_list[val] != lang_books[val]:
#             ws.Cells(terget, cell_idx).Value = lang_books[val]
#             kr_list[val] = lang_books[val]
#         else:
#             ws.Cells(terget, cell_idx).Value = kr_list[val]
#         ws.Cells(terget, cell_idx).Interior.ColorIndex = self.TABLE_CELL_COLOR # 색상적용
#         self.set_Win32com_cellStyle(ws = ws, terget = terget, cell_idx = cell_idx, heightSize = None)
#         print(f'time : {val}')

#     # 엑셀 데이터 적용
#     for key_terget in range(0, len(self.create_imgCellCount(lang=lang))):
    
#         cellCount = key_terget + 2
#         if cellCount <= len(self.create_imgCellCount(lang=lang)):
#             try:
#                 path = self.create_imgCellCount(lang=lang)[key_terget]
#                 img = IMG(key_terget)
#                 img.width = self.IMG_WIDTHSIZE
#                 img.height = self.IMG_HEIGHTSIZE
            
#                 image = ws.Shapes.AddPicture(path, False, True, cellCount, 1, img.width, img.height)
#                 ws.Cells(cellCount, 1).Value = path
#             except:
#                 ws.Cells(cellCount, 1).Value = "경로 없음"
#             time.sleep(0.005)

#             data_List = list(self.overloading_cell_search_val(sequence= key_terget, lang= lang))

#             for val in range(0, len(kr_list)):
#                 if kr_list[val] != data_List[val]:
#                     cell_idx = val + 1
#                     ws.Cells(cellCount, cell_idx).Value = ("" if "" == data_List[val] else data_List[val])
#                     self.set_Win32com_cellStyle(ws = ws, terget = cellCount, cell_idx = cell_idx, heightSize = self.IMG_SHEET_HEIGHTSIZE)
#             print(f'진행도 : {cellCount}/{len(self.create_imgCellCount(lang=lang))}')
#             time.sleep(0.005)
#             QApplication.processEvents()
            
#     def create_sheet_history(self, lang:str, ws:object):
#         """
#         연속적으로 생성 할 수 있는 규칙 필요
#         """

#         lang_books = []
#         lang_books.append("언어")

        
#         cell_idx = 1
#         cell_row = self.history_rows
#         for val in self.create_tupleBooks(lang=lang):
#             lang_books.append(val)

#         # 컬럼 설정
#         for val in range(0, len(lang_books)):
#             ws.cell(row = self.history_rows, # 현재 진행상황
#                         column= cell_idx, # 평가한 나라 개수 영향
#                         value=lang_books[val])
#             cell_idx = cell_idx + 1

#         # 언어 이미지명 testList Field 버전
        
#         sequence = 0
#         self.history_rows = self.history_rows + 1
#         for path in self.create_imgCellCount(lang = lang):

#             if (path != ""):
                
#                 print(f'terget : {self.overloading_cell_search_val(sequence = sequence, lang= lang)}')
#                 if "FAIL" in self.overloading_cell_search_val(sequence = sequence, lang= lang) or\
#                     "N/A" in self.overloading_cell_search_val(sequence = sequence, lang= lang) or\
#                         "N/T" in self.overloading_cell_search_val(sequence = sequence, lang= lang):

#                     columns= 0 + 1
#                     print(f'columns : {columns}')
#                     print(f'Value : {lang}')
#                     ws.cell(row = self.history_rows,
#                             column= columns,
#                             value = lang)

#                     columns = columns + 1
#                     columnsVal = 0
#                     print(f'columns : {columns}')
#                     print(f'Value : {self.cell_search_val(sequence = sequence, columns = columnsVal, lang = lang)}')
#                     ws.cell(row = self.history_rows,
#                             column= columns,
#                             value = self.cell_search_val(sequence = sequence, columns = columnsVal, lang = lang))
                    
#                     # 평가 결과
#                     for idx in range(0, len(self.evaluation_len(key="Test_List"))):

#                         columns= columns + 1
#                         columnsVal = columnsVal + 1
#                         print(f'columns : {columns}')
#                         print(f'Value : {self.cell_search_val(sequence = sequence, columns = columnsVal, lang = lang)}')
#                         value = self.cell_search_val(sequence = sequence, columns = columnsVal, lang = lang)

#                         ws.cell(row = self.history_rows,
#                             column= columns,
#                             value = ("" if "" == value else value))

#                     # 라벨 평가
#                     for idx in range(0, len(self.evaluation_len(key="Field"))):

#                         columns= columns + 1
#                         columnsVal = columnsVal + 1
#                         print(f'columns : {columns}')
#                         print(f'Value : {self.cell_search_val(sequence = sequence, columns = columnsVal, lang = lang)}')
#                         value = self.cell_search_val(sequence = sequence, columns = columnsVal, lang = lang)

#                         ws.cell(row = self.history_rows,
#                             column= columns,
#                             value = ("" if "" == value else value))

                
#                     # 버전 정보
#                     print(f'columns : {columns + 1}')
#                     print(f'Value : {self.cell_search_val(sequence = sequence, columns = columnsVal + 1, lang = lang)}')
#                     value = self.cell_search_val(sequence = sequence, columns = columnsVal + 1, lang = lang)
#                     ws.cell(row = self.history_rows,
#                                 column= columns + 1,
#                                 value = "" if "" == value else value)

#                     self.history_rows = self.history_rows + 1
#                 sequence = sequence + 1

#         self.set_cellStyle(active= ws, idx = cell_row, fix= "SUMMARY")
#         self.history_rows = self.history_rows + 1

#     def update_sheet_history(self, lang:str, ws:object, integer:int):

#         row = self.historyUpdate_rows
#         colCount = 1
#         colList = []
#         # 새로운 데이터 추가 부여필요
#         tuples = []
#         tuples.append("언어")
#         for tuple in self.create_tupleBooks(lang= lang):
#             tuples.append(tuple)

#         self.set_Win32com_cellStyle(ws = ws, terget = row - 1, cell_idx = colCount, heightSize = None)
#         while(ws.Cells(row, integer).Value == lang):
#             print(f"확인 : {ws.Cells(row, integer).Value}")
#             row = row + 1

#         while(ws.Cells(self.historyUpdate_rows - 1, colCount).Value != None):
#             print(f'colCountValue : {ws.Cells(integer, colCount).Value}')
#             colList.append(ws.Cells(self.historyUpdate_rows - 1, colCount).Value)
#             colCount = colCount + 1
#         colCount = colCount - 1

#         print(f'colCount : {colCount}')
#         print(f'lang : {lang}')
#         print(f'DB_colCount : {len(self.create_tupleBooks(lang= lang))}')
#         # 컬럼개수의 변동이 있었는가?
#         if (colList == tuples):
#             getPathList = []
#             pathList = []
            
#             for row_idx in range(self.historyUpdate_rows, row):
#                 print(f"ws.Cells(row_idx, integer).Value : {ws.Cells(row_idx, integer).Value}")
#                 if (ws.Cells(row_idx, integer).Value == lang):
#                     getPathList.append(os.path.basename(ws.Cells(row_idx, 2).Value)) # 파일명만 추출

#             for idx in range(0, len(self.create_imgCellCount(lang = lang))):

#                 data_List = list(self.overloading_cell_search_val(sequence= idx, lang= lang))
#                 if ("FAIL" in data_List or "N/A" in data_List or "N/T" in data_List):
#                     pathList.append(os.path.basename(self.create_imgCellCount(lang)[idx])) # 파일명만 추출모음
                
#             print(f'getPathList : {len(getPathList)}')
#             print(f'pathList : {len(pathList)}')

#             if len(getPathList) != len(pathList):

#                 print(f'pathList : {len(pathList)}')
#                 print(f'getPathList : {len(getPathList)}')

#                 char_lenght = len(pathList) - len(getPathList)
#                 print(f'char_lenght : {char_lenght}')
#                 upper = [f'{i}' for i in string.ascii_uppercase]
#                 if (char_lenght > 0):
                    
#                     new_style_terget = f"{upper[integer - 1]}{self.historyUpdate_rows}:{upper[len(self.create_tupleBooks(lang = lang))]}{self.historyUpdate_rows}"
#                     for i in range(0, char_lenght):
#                         ws.Rows(self.historyUpdate_rows).Insert()
#                         ws.Rows(self.historyUpdate_rows).RowHeight = self.IMG_SHEET_HEIGHTSIZE
#                         ws.Range(f"{new_style_terget}").Interior.ColorIndex = 2
#                         ws.Cells(self.historyUpdate_rows, integer).Value = lang
#                         pathList.append("NULL")
#                 elif (char_lenght < 0):
#                     for i in range(0, abs(char_lenght)):
#                         ws.Rows(self.historyUpdate_rows).Delete()
#                         del pathList[len(pathList) - 1]

#             for idx in range(0, len(pathList)):
            
#                 jdx = 2
#                 for value in self.overloading_cell_search_val(sequence= idx, lang= lang):

#                     if ("FAIL" in self.overloading_cell_search_val(sequence= idx, lang= lang) or\
#                         "N/A" in self.overloading_cell_search_val(sequence= idx, lang= lang) or\
#                             "N/T" in self.overloading_cell_search_val(sequence= idx, lang= lang)):
#                         ws.Cells(self.historyUpdate_rows, jdx).Value = value
#                         jdx= jdx + 1
#                 self.historyUpdate_rows = self.historyUpdate_rows + 1
            
#         else:
            
#             getPathList = []
#             pathList = []

#             for row_idx in range(self.historyUpdate_rows, row):
#                 print(f"ws.Cells(row_idx, integer).Value : {ws.Cells(row_idx, integer).Value}")
#                 if (ws.Cells(row_idx, integer).Value == lang):
#                     getPathList.append(os.path.basename(ws.Cells(row_idx, 2).Value)) # 파일명만 추출

#             for idx in range(0, len(self.create_imgCellCount(lang = lang))):

#                 data_List = list(self.overloading_cell_search_val(sequence= idx, lang= lang))
#                 if ("FAIL" in data_List or "N/A" in data_List or "N/T" in data_List):
#                     pathList.append(os.path.basename(self.create_imgCellCount(lang)[idx])) # 파일명만 추출모음
                
#             print(f'getPathList : {len(getPathList)}')
#             print(f'pathList : {len(pathList)}')

#             if len(getPathList) != len(pathList):

#                 print(f'pathList : {len(pathList)}')
#                 print(f'getPathList : {len(getPathList)}')

#                 char_lenght = len(pathList) - len(getPathList)
#                 print(f'char_lenght : {char_lenght}')
#                 if (char_lenght > 0):
                    
#                     for i in range(0, char_lenght):
#                         ws.Rows(row).Insert()
                        
#                 elif (char_lenght < 0):
#                     for i in range(0, abs(char_lenght)):
#                         ws.Rows(row).Delete()

#             row = row - 1
#             self.historyUpdate_rows = self.historyUpdate_rows - 1
#             print(f'row : {row}')
#             for i in range(0, row):
#                 ws.Rows(self.historyUpdate_rows).Delete()
            
#             for i in range(0, row):
#                 ws.Rows(self.historyUpdate_rows).Insert()

#             # 태그적용
#             for val in range(0, len(tuples)):
#                 ws.Cells(self.historyUpdate_rows, val + 1).Value = tuples[val]
#                 ws.Cells(self.historyUpdate_rows, val + 1).Interior.ColorIndex = self.TABLE_CELL_COLOR # 색상적용
#                 self.set_Win32com_cellStyle(ws = ws, terget = self.historyUpdate_rows, cell_idx = val + 1, heightSize = None)

#                 print(f'tuples[val] = {tuples[val]}')
#                 if tuples[val] == f"{lang}_이미지":
#                     print("이미지 사이즈 변경")
#                     ws.Range(f"A{val + 1}").Columns.Autofit
#                 elif tuples[val] in self.evaluation_len(key="Test_List") or tuples[val] == "언어":
#                     ws.Columns(val + 1).ColumnWidth = self.SHEET_WIDTH_SHORTSIZE  # 선택 영역 행 크기 설정
                


#             # 엑셀 데이터 적용 NA NT FAIL 분류하기
#             for key_terget in range(0, len(self.create_imgCellCount(lang=lang))):

#                 data_List = list(self.overloading_cell_search_val(sequence= key_terget, lang= lang))
#                 if ("FAIL" in data_List or "N/A" in data_List or "N/T" in data_List):
#                     self.historyUpdate_rows = self.historyUpdate_rows + 1
#                     resultList = []
#                     resultList.append(lang)
#                     for key in data_List:
#                         resultList.append(key)

#                     for val in range(0, len(resultList)):
#                         ws.Cells(self.historyUpdate_rows, val + 1).Value = ("" if "" == resultList[val] else resultList[val])
#                         self.set_Win32com_cellStyle(ws = ws, terget = self.historyUpdate_rows, cell_idx = val + 1, heightSize = self.IMG_SHEET_HEIGHTSIZE)
                    
#                 print(f'SUMMARY 진행도 : {key_terget}/{len(self.create_imgCellCount(lang=lang))}')
#                 time.sleep(0.005)
#                 QApplication.processEvents()

#             count = 1
#             while(True):

#                 self.historyUpdate_rows = self.historyUpdate_rows + 1        
#                 print(f"확인 : {ws.Cells(self.historyUpdate_rows, integer).Value}")
#                 if ws.Cells(self.historyUpdate_rows, integer).Value == lang or ws.Cells(self.historyUpdate_rows, integer).Value == "언어":
#                     break
#                 elif count == 10:
#                     print("카운트 멈춤")
#                     break
#                 elif ws.Cells(self.historyUpdate_rows, integer).Value == None:
#                     print(f"카운트 적립 : {count}")
#                     count = count + 1
#                 else:
#                     ws.Rows(self.historyUpdate_rows).Delete()

#             ws.Rows(self.historyUpdate_rows).Insert()
#         self.historyUpdate_rows = self.historyUpdate_rows + 1       


                
                
                

